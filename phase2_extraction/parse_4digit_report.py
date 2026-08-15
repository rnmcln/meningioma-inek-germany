#!/usr/bin/env python3
"""
parse_4digit_report.py — extract 4-digit meningeal codes (D32.0/.1/.9, D42.0,
C70.0) from a Destatis "Statistischer Bericht - Diagnosen der Krankenhauspatienten"
Excel (the 4-digit source the GENESIS 23131 cube does not provide).

Usage:
  python parse_4digit_report.py [file.xlsx]     # defaults to newest matching file

Pass 1 (structure unknown): scans every sheet for the target codes and prints the
matching rows with context, so the exact extraction rule can be finalised.
Writes a tidy CSV of whatever total-column values it can identify.
"""
import sys, os, glob, csv, re

TARGETS = ["D32.0", "D32.1", "D32.9", "D42.0", "C70.0"]

def find_default(here):
    pats = glob.glob(os.path.join(here, "data", "*iagnose*.xlsx")) + \
           glob.glob(os.path.join(here, "data", "*5231301*.xlsx")) + \
           glob.glob(os.path.join(os.path.dirname(here), "*5231301*.xlsx")) + \
           glob.glob(os.path.join(os.path.dirname(here), "*iagnose*.xlsx"))
    return max(pats, key=os.path.getmtime) if pats else None

def main():
    here = os.path.dirname(os.path.abspath(__file__))
    path = sys.argv[1] if len(sys.argv) > 1 else find_default(here)
    if not path or not os.path.exists(path):
        sys.exit("No Excel found. Pass the .xlsx path, or drop it in phase2_extraction/data/.")
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl required: pip install openpyxl --break-system-packages")

    print(f"Reading: {path}")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    print(f"Sheets: {wb.sheetnames}\n")

    hits = []
    for ws in wb.worksheets:
        for ri, row in enumerate(ws.iter_rows(values_only=True), 1):
            cells = [("" if c is None else str(c)) for c in row]
            joined = " | ".join(cells)
            for code in TARGETS:
                # match D32.0 as a token (Destatis may write 'D32.0' or 'D320' or 'D32.-')
                if re.search(rf"\b{re.escape(code)}\b", joined) or code.replace(".", "") in joined.replace(".", ""):
                    nums = [c for c in cells if re.fullmatch(r"-?\d[\d.]*", c.strip())]
                    hits.append((ws.title, ri, code, joined[:200], nums[:6]))
                    break
    if not hits:
        print("No target codes found. Inspect sheet layout manually; codes may be split "
              "across columns (code in one cell, label in another).")
        # dump a sample of the first data sheet for inspection
        ws = wb.worksheets[0]
        for ri, row in enumerate(ws.iter_rows(values_only=True), 1):
            if ri > 40: break
            print(ri, [("" if c is None else str(c))[:22] for c in row][:8])
        return

    print(f"Found {len(hits)} matching rows:\n")
    for sheet, ri, code, ctx, nums in hits[:60]:
        print(f"[{sheet} r{ri}] {code}: {ctx}")
        print(f"      numeric cells: {nums}")

    out = os.path.join(here, "data", "destatis_4digit_meningeal_RAW.csv")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["sheet", "row", "code", "context", "numeric_cells"])
        for h in hits:
            w.writerow([h[0], h[1], h[2], h[3], " ; ".join(h[4])])
    print(f"\nWrote raw matches: {out}")
    print("Next: confirm which numeric column is the total (all ages, both sexes) "
          "and I will finalise the tidy extraction + reconciliation update.")

if __name__ == "__main__":
    main()
